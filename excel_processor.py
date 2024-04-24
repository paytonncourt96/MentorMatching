import streamlit as st
import pandas as pd
import base64
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to process existing Excel files
def process_excel(file):
    df = pd.read_excel(file)
    if 'Name' in df.columns and 'Email' in df.columns:
        return df[['Name', 'Email']]
    else:
        st.error("Excel file must contain 'Name' and 'Email' columns.")

# Function to extract submissions from full names
def extract_submissions(full_name):
    if pd.isna(full_name):
        return ""
    parts = full_name.split()
    return parts[0] + ' ' + parts[1][0] if len(parts) > 1 else full_name

# Function to normalize skill names
def normalize_skill_name(skill):
    return skill.lower().strip()

# Function to map skills to target DataFrame
def map_skills_to_target(row, target_columns_normalized):
    target_values = {column: None for column in target_columns_normalized}

    skill_column_names = [
        'Select your first choice of skill from the list below',
        'Select your second choice of skill from the list below',
        'Select your third choice of skill from the list below',
        'Select your fourth choice of skill from the list below (if applicable)',
        'Select your fifth choice of skill from the list below (if applicable)'
    ]

    for idx, skill_column_name in enumerate(skill_column_names):
        skill_key = normalize_skill_name(row[skill_column_name]) if not pd.isna(row[skill_column_name]) else ''
        if skill_key in target_columns_normalized:
            target_values[skill_key] = idx + 1

    return target_values

# Function to apply background color based on values
def apply_color_based_on_value(ws, start_row, start_col, end_col, value_to_color_map):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if cell.value in value_to_color_map:
                cell.fill = value_to_color_map[cell.value]

# Main function
def main():
    st.set_page_config(page_title="Mentor Matching App", page_icon="IUF_logo_white.png") 
    logo_image = 'IUF_logo_black.png'

    st.image(logo_image, use_column_width=True)
    st.title("Mentor Matching")
    st.subheader("Upload an Excel file and extract Names and Emails")

    # First file uploader and processing
    existing_file = st.file_uploader("Upload Existing Excel file", type=['xlsx', 'xls'], accept_multiple_files=False)
    
    if existing_file is not None:
        st.write("Uploaded file:", existing_file.name)
        
        df_existing = process_excel(existing_file)
        if df_existing is not None:
            st.write("Extracted Names and Emails:")
            st.write(df_existing)

            csv_existing = df_existing.to_csv(index=False)
            b64_existing = base64.b64encode(csv_existing.encode()).decode()
            href_existing = f'<a href="data:file/csv;base64,{b64_existing}" download="output.csv">Download CSV File</a>'
            st.markdown(href_existing, unsafe_allow_html=True)

    # Second file uploader and processing
    new_file = st.file_uploader("Upload New Excel file", type=['xlsx', 'xls'], accept_multiple_files=False)
    
    if new_file is not None:
        st.write("Uploaded file:", new_file.name)

        # Load the template workbook
        template_url = 'Template.xlsx'
        wb_template = load_workbook(template_url)
        ws_template = wb_template.active

        target_columns = [
            'analytical thinking', 'business processes', 'decision making', 'effective communication / listening',
            'negotiation', 'managing change', 'data analytics / literacy', 'problem solving', 'managing resources',
            'project management', 'conflict management', 'using financials', 'presentations',
            'collaborating with others', 'compliance practices', 'legal considerations', 'fundraising principles',
            'real estate practices', 'policies / procedures principles', 'customer service', 'facilitation',
            'branding & marketing', 'business communications', 'planning & organizing', 'administrative practices',
            'building relationships', 'systems design & thinking', 'navigating cultural differences',
            'navigating organizational structures', 'technology incorporation', 'database management',
            'accounting operations skills', 'investments operations skills'
        ]
        
        target_columns_normalized = [normalize_skill_name(col) for col in target_columns]

        wb_new = load_workbook(new_file)
        ws_new = wb_new.active

        df = pd.DataFrame(columns=[
            'Application Type', 'Submissions', 'Application Date'] + target_columns_normalized)

        for index, row in pd.read_excel(new_file).iterrows():
            application_type = row['Are you interested in being a mentor or mentee?']
            submissions = extract_submissions(row['Name'])
            application_date = pd.to_datetime(row['Completion time']).strftime('%m/%d/%Y')
            skills_values = map_skills_to_target(row, target_columns_normalized)

            df = df.append({
                'Application Type': application_type,
                'Submissions': submissions,
                'Application Date': application_date,
                **skills_values
            }, ignore_index=True)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
            for c_idx, value in enumerate(row, 1):
                ws_new.cell(row=r_idx, column=c_idx, value=value)

        value_to_color_map = {
            1: PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),  # Green
            2: PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
            3: PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # Orange
            4: PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid'),  # Pink
            5: PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid'),  # Light Blue
        }

        apply_color_based_on_value(ws_new, 2, 4, ws_new.max_column, value_to_color_map)

        # Display the download button for the new Excel file
        st.write("Processed Excel file:")
        st.write(df)

        new_excel_data = io.BytesIO()
        wb_new.save(new_excel_data)
        b64_new = base64.b64encode(new_excel_data.getvalue()).decode()
        href_new = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_new}" download="PairingResults.xlsx">Download Processed Excel File</a>'
        st.markdown(href_new, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
